
from django.contrib.auth.models import User
from django.http import HttpResponse

from openpyxl.styles import PatternFill, Font
from openpyxl.styles import Alignment
from openpyxl import Workbook

from dateutil.relativedelta import relativedelta
from datetime import date, datetime

import calendar

from budgetingApp.models import Finances
from budgetingApp.models import UserSettings

def export_excel(user: User, month_year: date):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Finances.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Financies"

    # headers
    headers = ["Type", "Category", "Amount", "Currency", "Date"]
    ws.append(headers)

    # colors to use
    headerFill = PatternFill(start_color='4d4d4d',
                        end_color='4d4d4d',
                        fill_type='solid')
    incomeFill = PatternFill(start_color='e6e6e6',
                        end_color='e6e6e6',
                        fill_type='solid')
    expenseFill = PatternFill(start_color='cccccc',
                        end_color='cccccc',
                        fill_type='solid')

    # number of rows added
    no_row = 1

    # adding data from database
    if month_year != None:
        today = month_year
    else:
        today = date.today()
        today = today + relativedelta(months=-1)

    month = datetime.strptime(today, "%Y-%m").month
    year = datetime.strptime(today, "%Y-%m").year

    sumIncome = {}
    sumExpense = {}
    remaining = {}

    MainCurrency = UserSettings.getObjectsByUser(user)[0]

    for obj in Finances.getObjectsByUser(user):
        # adding incomes for current month
        if month == obj.Date.month and year == obj.Date.year :
            if obj.Type == 0:
                ws.append(["Income", obj.Category, "+" + str(obj.Amount), obj.Currency, str(obj.Date)])
                no_row = no_row + 1

                ws['A' + str(no_row)].fill = incomeFill
                ws['B' + str(no_row)].fill = incomeFill
                ws['C' + str(no_row)].fill = incomeFill
                ws['D' + str(no_row)].fill = incomeFill
                ws['E' + str(no_row)].fill = incomeFill
                ws.cell(row = no_row, column = 1).font = Font(size = 12, italic = True)
                ws.cell(row = no_row, column = 2).font = Font(size = 12)
                ws.cell(row = no_row, column = 3).font = Font(size = 12)
                ws.cell(row = no_row, column = 4).font = Font(size = 12)
                ws.cell(row = no_row, column = 5).font = Font(size = 12)

                try:
                    sumIncome[MainCurrency] += obj.Amount
                    remaining[obj.Currency] += obj.Amount
                except KeyError:
                    sumIncome[MainCurrency] = obj.Amount
                    remaining[MainCurrency] = obj.Amount

            elif obj.Type == 1: #adding expenses for current month
                ws.append(["Expense", obj.Category, "-" + str(obj.Amount), obj.Currency, str(obj.Date)])
                no_row = no_row + 1

                ws['A' + str(no_row)].fill = expenseFill
                ws['B' + str(no_row)].fill = expenseFill
                ws['C' + str(no_row)].fill = expenseFill
                ws['D' + str(no_row)].fill = expenseFill
                ws['E' + str(no_row)].fill = expenseFill

                ws.cell(row = no_row, column = 1).font = Font(size = 12, italic = True)
                ws.cell(row = no_row, column = 2).font = Font(size = 12)
                ws.cell(row = no_row, column = 3).font = Font(size = 12)
                ws.cell(row = no_row, column = 4).font = Font(size = 12)
                ws.cell(row = no_row, column = 5).font = Font(size = 12)

                try:
                    sumExpense[MainCurrency] += obj.Amount
                except KeyError:
                    sumExpense[MainCurrency] = obj.Amount
                    
                try:
                    remaining[MainCurrency] -= obj.Amount
                except KeyError:
                    remaining[MainCurrency] = -obj.Amount

            currentCell = ws['C' + str(no_row)] 
            currentCell.alignment = Alignment(horizontal='right')

            currentCell = ws['D' + str(no_row)] 
            currentCell.alignment = Alignment(horizontal='center')

            currentCell = ws['E' + str(no_row)] 
            currentCell.alignment = Alignment(horizontal='center')
  

    # adding and formating lines with totals
    ws.cell(row = no_row + 2, column = 2).value = "Total income"
    ws.cell(row = no_row + 2, column = 3).value = "+" + str(sumIncome[MainCurrency])
    ws.cell(row = no_row + 2, column = 3).alignment = Alignment(horizontal='right')

    ws.cell(row = no_row + 3, column = 2).value = "Total expenses"
    ws.cell(row = no_row + 3, column = 3).value = "-" + str(sumExpense[MainCurrency])
    ws.cell(row = no_row + 3, column = 3).alignment = Alignment(horizontal='right')

    ws.cell(row = no_row + 4, column = 2).value = "Remaining money"
    ws.cell(row = no_row + 4, column = 3).value = str(remaining[MainCurrency])
    ws.cell(row = no_row + 4, column = 3).alignment = Alignment(horizontal='right')

    ws.cell(row = no_row + 2, column = 1).fill = incomeFill
    ws.cell(row = no_row + 2, column = 2).fill = incomeFill
    ws.cell(row = no_row + 2, column = 3).fill = incomeFill
    ws.cell(row = no_row + 2, column = 4).fill = incomeFill
    ws.cell(row = no_row + 2, column = 5).fill = incomeFill

    ws.cell(row = no_row + 3, column = 1).fill = expenseFill
    ws.cell(row = no_row + 3, column = 2).fill = expenseFill
    ws.cell(row = no_row + 3, column = 3).fill = expenseFill
    ws.cell(row = no_row + 3, column = 4).fill = expenseFill
    ws.cell(row = no_row + 3, column = 5).fill = expenseFill

    ws.cell(row = no_row + 4, column = 1).fill = incomeFill
    ws.cell(row = no_row + 4, column = 2).fill = incomeFill
    ws.cell(row = no_row + 4, column = 3).fill = incomeFill
    ws.cell(row = no_row + 4, column = 4).fill = incomeFill
    ws.cell(row = no_row + 4, column = 5).fill = incomeFill

    ws.cell(row = no_row + 1, column = 1).fill = headerFill
    ws.cell(row = no_row + 1, column = 2).fill = headerFill
    ws.cell(row = no_row + 1, column = 3).fill = headerFill
    ws.cell(row = no_row + 1, column = 4).fill = headerFill
    ws.cell(row = no_row + 1, column = 5).fill = headerFill

    
    # formating the header
    ws.row_dimensions[1].height = 20
    ws['A1'].fill = headerFill
    ws['B1'].fill = headerFill
    ws['C1'].fill = headerFill
    ws['D1'].fill = headerFill
    ws['E1'].fill = headerFill

    currentCell = ws['E1'] 
    currentCell.alignment = Alignment(horizontal='center')
    
    ws.column_dimensions['A'].width = 12
    ws.cell(row = 1, column = 1).font = Font(size = 14, bold = True, color = 'FFFFFFFF') 
    ws.column_dimensions['B'].width = 18
    ws.cell(row = 1, column = 2).font = Font(size = 14, bold = True, color = 'FFFFFFFF') 
    ws.column_dimensions['C'].width = 12
    ws.cell(row = 1, column = 3).font = Font(size = 14, bold = True, color = 'FFFFFFFF') 
    ws.column_dimensions['D'].width = 12
    ws.cell(row = 1, column = 4).font = Font(size = 14, bold = True, color = 'FFFFFFFF') 
    ws.column_dimensions['E'].width = 14
    ws.cell(row = 1, column = 5).font = Font(size = 14, bold = True, color = 'FFFFFFFF') 

    # saving everything
    wb.save(response)
    return response
